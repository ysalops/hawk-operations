import os
import unicodedata
import hashlib
import hmac
import base64
import csv
import io
import json
import re

from datetime import date, datetime
from pathlib import Path

from fastapi import Depends, FastAPI, File, Form, HTTPException, Request, UploadFile, status
from fastapi.responses import FileResponse, JSONResponse, RedirectResponse
from pydantic import BaseModel
from fastapi.staticfiles import StaticFiles
from sqlalchemy import inspect, or_, select, text
from sqlalchemy.orm import Session

from backend import models, schemas
from backend.database import Base, engine, get_db


# BANCO DE DADOS

Base.metadata.create_all(bind=engine)


def aplicar_migracoes_leves() -> None:
    """Adiciona campos opcionais sem apagar dados existentes."""
    inspetor = inspect(engine)

    if "motoristas" not in inspetor.get_table_names():
        return

    colunas = {
        coluna["name"]
        for coluna in inspetor.get_columns("motoristas")
    }

    alteracoes = {
        "cpf": "VARCHAR(11)",
        "cnh": "VARCHAR(30)",
        "categoria_cnh": "VARCHAR(10)",
        "validade_cnh": "DATE",
    }

    with engine.begin() as conexao:
        for nome_coluna, tipo_sql in alteracoes.items():
            if nome_coluna not in colunas:
                conexao.execute(
                    text(
                        f"ALTER TABLE motoristas ADD COLUMN {nome_coluna} {tipo_sql}"
                    )
                )

        conexao.execute(
            text(
                "CREATE UNIQUE INDEX IF NOT EXISTS "
                "ux_motoristas_cpf ON motoristas (cpf)"
            )
        )

    inspetor = inspect(engine)
    if "operacoes" in inspetor.get_table_names():
        colunas_operacoes = {
            coluna["name"]
            for coluna in inspetor.get_columns("operacoes")
        }
        with engine.begin() as conexao:
            if "ajudante_id" not in colunas_operacoes:
                conexao.execute(
                    text("ALTER TABLE operacoes ADD COLUMN ajudante_id INTEGER")
                )
            conexao.execute(
                text(
                    "CREATE INDEX IF NOT EXISTS "
                    "ix_operacoes_ajudante_id ON operacoes (ajudante_id)"
                )
            )


aplicar_migracoes_leves()


# CAMINHOS

BASE_DIR = Path(__file__).resolve().parent.parent
FRONTEND_DIR = BASE_DIR / "frontend"

NO_CACHE_HEADERS = {
    "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
    "Pragma": "no-cache",
    "Expires": "0",
}

# AUTENTICAÇÃO

YLUME_OPS_ACCESS_PASSWORD = os.getenv(
    "YLUME_OPS_ACCESS_PASSWORD",
    "",
).strip()

YLUME_OPS_SESSION_SECRET = os.getenv(
    "YLUME_OPS_SESSION_SECRET",
    "",
).strip()

SESSION_COOKIE_NAME = "ylume_ops_session"
SESSION_MAX_AGE = 60 * 60 * 8  # 8 horas

# IMPORTAÇÃO INTELIGENTE
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "").strip()
YLUME_OPS_AI_MODEL = os.getenv("YLUME_OPS_AI_MODEL", "gpt-5.6-luna").strip() or "gpt-5.6-luna"
MAX_IMPORT_FILE_BYTES = 10 * 1024 * 1024
MAX_IMPORT_IMAGES = 4

YLUME_OPS_COOKIE_SECURE = (
    os.getenv(
        "YLUME_OPS_COOKIE_SECURE",
        "true",
    )
    .strip()
    .lower()
    in {
        "1",
        "true",
        "yes",
        "on",
    }
)

def criar_token_sessao() -> str:
    if not YLUME_OPS_SESSION_SECRET:
        return ""

    return hmac.new(
        YLUME_OPS_SESSION_SECRET.encode("utf-8"),
        b"ylume-ops-session-v1",
        hashlib.sha256,
    ).hexdigest()


def esta_autenticado(request: Request) -> bool:
    token = request.cookies.get(
        SESSION_COOKIE_NAME,
    )

    token_esperado = criar_token_sessao()

    if not token or not token_esperado:
        return False

    return hmac.compare_digest(
        token,
        token_esperado,
    )


class LoginRequest(BaseModel):
    password: str

# APLICAÇÃO

app = FastAPI(
    title="Ylume Ops API",
    description=(
        "API para gestão operacional de frota, motoristas, ajudantes, "
        "manutenções, panoramas e rotas."
    ),
    version="1.0.1",
)

@app.middleware("http")
async def proteger_ylume_ops(
    request: Request,
    call_next,
):
    caminho = request.url.path

    rotas_publicas = {
        "/login",
        "/auth/login",
        "/health",
    }

    if (
        caminho in rotas_publicas
        or caminho.startswith("/static/")
    ):
        return await call_next(request)

    if esta_autenticado(request):
        return await call_next(request)

    if request.method == "GET":
        return RedirectResponse(
            url="/login?v=1.0.1",
            status_code=303,
        )

    return JSONResponse(
        status_code=401,
        content={
            "detail": "Sessão não autenticada.",
        },
    )

# ARQUIVOS ESTÁTICOS

app.mount(
    "/static",
    StaticFiles(directory=FRONTEND_DIR),
    name="static",
)


# HOME / HEALTH

@app.get(
    "/login",
    include_in_schema=False,
)
def login_page(
    request: Request,
):
    if esta_autenticado(request):
        return RedirectResponse(
            url="/?v=1.0.1",
            status_code=303,
        )

    return FileResponse(
        FRONTEND_DIR / "login.html",
        headers=NO_CACHE_HEADERS,
    )


@app.post(
    "/auth/login",
    include_in_schema=False,
)
def login(
    dados: LoginRequest,
):
    if (
        not YLUME_OPS_ACCESS_PASSWORD
        or not YLUME_OPS_SESSION_SECRET
    ):
        raise HTTPException(
            status_code=503,
            detail=(
                "Autenticação do Ylume Ops "
                "não foi configurada."
            ),
        )

    if not hmac.compare_digest(
        dados.password,
        YLUME_OPS_ACCESS_PASSWORD,
    ):
        raise HTTPException(
            status_code=401,
            detail="Senha inválida.",
        )

    response = JSONResponse(
        content={
            "authenticated": True,
        },
    )

    response.set_cookie(
        key=SESSION_COOKIE_NAME,
        value=criar_token_sessao(),
        max_age=SESSION_MAX_AGE,
        httponly=True,
        secure=YLUME_OPS_COOKIE_SECURE,
        samesite="lax",
    )

    return response


@app.post(
    "/auth/logout",
    include_in_schema=False,
)
def logout():
    response = JSONResponse(
        content={
            "authenticated": False,
        },
    )

    response.delete_cookie(
        SESSION_COOKIE_NAME,
    )

    return response


@app.get("/", include_in_schema=False)
def home():
    return FileResponse(
        FRONTEND_DIR / "index.html",
        headers=NO_CACHE_HEADERS,
    )


@app.get("/health")
def health_check():
    return {"status": "ok"}


# FUNÇÕES AUXILIARES

def limpar_texto_opcional(valor: str | None) -> str | None:
    if valor is None:
        return None

    valor_limpo = valor.strip()
    return valor_limpo or None


def normalizar_cpf(valor: str | None) -> str | None:
    if valor is None:
        return None

    cpf = "".join(
        caractere
        for caractere in valor
        if caractere.isdigit()
    )

    if not cpf:
        return None

    if len(cpf) != 11 or len(set(cpf)) == 1:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Informe um CPF válido com 11 dígitos.",
        )

    for tamanho in (9, 10):
        soma = sum(
            int(cpf[indice]) * (tamanho + 1 - indice)
            for indice in range(tamanho)
        )
        digito = (soma * 10) % 11
        if digito == 10:
            digito = 0

        if digito != int(cpf[tamanho]):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Informe um CPF válido.",
            )

    return cpf


def criar_chave_busca(valor: str | None) -> str | None:
    texto = limpar_texto_opcional(valor)

    if not texto:
        return None

    texto_sem_acento = "".join(
        caractere
        for caractere in unicodedata.normalize(
            "NFKD",
            texto,
        )
        if not unicodedata.combining(
            caractere
        )
    )

    return " ".join(
        texto_sem_acento
        .casefold()
        .split()
    )



def combinar_textos_operacao(
    cluster: str | None,
    observacao: str | None,
) -> str | None:
    partes = []

    for valor in (
        cluster,
        observacao,
    ):
        texto = limpar_texto_opcional(
            valor
        )

        if (
            texto
            and
            texto not in partes
        ):
            partes.append(
                texto
            )

    return (
        " | ".join(
            partes
        )
        or
        None
    )



# =====================================================
# IMPORTAÇÃO INTELIGENTE - FUNÇÕES AUXILIARES
# =====================================================

STATUS_IMPORTACAO_INTELIGENTE = {
    "CARREGANDO",
    "EM_ROTA",
    "CONCLUIDA",
    "RETORNANDO_ESTACAO",
    "AMBULANCIA",
    "RESERVA_CARREGANDO",
    "FOLGA",
    "IMPEDIDO",
    "SEM_CARGA",
    "OUTRO_SERVICE",
    "INDISPONIVEL_MOTORISTA",
    "SEM_CLASSIFICACAO",
    "MANUTENCAO",
}


def normalizar_placa_importacao(valor: str | None) -> str:
    if not valor:
        return ""
    return re.sub(r"[^A-Za-z0-9]", "", valor).upper()


def normalizar_status_importacao(valor: str | None) -> str:
    chave = criar_chave_busca(valor) or ""
    mapa = {
        "carregando": "CARREGANDO",
        "em rota": "EM_ROTA",
        "concluida": "CONCLUIDA",
        "concluido": "CONCLUIDA",
        "retornando a estacao": "RETORNANDO_ESTACAO",
        "ambulancia": "AMBULANCIA",
        "carro reserva": "RESERVA_CARREGANDO",
        "reserva carregando": "RESERVA_CARREGANDO",
        "folga": "FOLGA",
        "folga planejada motorista": "FOLGA",
        "impedido": "IMPEDIDO",
        "impedido de rodar": "IMPEDIDO",
        "sem carga": "SEM_CARGA",
        "outro service": "OUTRO_SERVICE",
        "rodou em outro service": "OUTRO_SERVICE",
        "indisponivel motorista": "INDISPONIVEL_MOTORISTA",
        "sem drive": "INDISPONIVEL_MOTORISTA",
        "manutencao": "MANUTENCAO",
        "sem classificacao": "SEM_CLASSIFICACAO",
    }
    if chave in mapa:
        return mapa[chave]
    candidato = (valor or "").strip().upper().replace(" ", "_")
    return candidato if candidato in STATUS_IMPORTACAO_INTELIGENTE else "SEM_CLASSIFICACAO"


def detectar_status_linha(linha: str, secao_manutencao: bool = False) -> tuple[str, float]:
    chave = criar_chave_busca(linha) or ""
    if secao_manutencao or "🛠" in linha or "manutencao" in chave:
        return "MANUTENCAO", 0.99
    if "🚗" in linha or "carro reserva" in chave:
        return "RESERVA_CARREGANDO", 0.98
    if "⚠" in linha or "folga" in chave:
        return "FOLGA", 0.98
    if "sem drive" in chave or "indisponivel" in chave:
        return "INDISPONIVEL_MOTORISTA", 0.98
    if "🚫" in linha or "impedido" in chave or "pausado" in chave:
        return "IMPEDIDO", 0.97
    if "📦" in linha or "sem carga" in chave:
        return "SEM_CARGA", 0.98
    if "🔄" in linha or "outro service" in chave:
        return "OUTRO_SERVICE", 0.98
    if "✅" in linha or "carregando" in chave:
        return "CARREGANDO", 0.98
    if "em rota" in chave:
        return "EM_ROTA", 0.92
    if "concluid" in chave:
        return "CONCLUIDA", 0.92
    return "SEM_CLASSIFICACAO", 0.65


def extrair_motivo_linha(linha: str, placa: str, rota_id: str | None, motorista: str | None) -> str | None:
    texto = linha
    texto = re.sub(re.escape(placa), " ", texto, flags=re.IGNORECASE)
    if rota_id:
        texto = texto.replace(rota_id, " ")
    if motorista:
        texto = re.sub(r"\([^)]*\)", " ", texto, count=1)
    for simbolo in ("✅", "🚗", "⚠️", "⚠", "🛠️", "🛠", "🚫", "📦", "🔄", "*"):
        texto = texto.replace(simbolo, " ")
    texto = re.sub(r"\s+", " ", texto).strip(" -–—.;")
    # Rótulos que já estão representados pelo status não precisam virar motivo.
    chave = criar_chave_busca(texto) or ""
    if chave in {"carregando", "folga", "sem carga", "sem drive", "manutencao"}:
        return None
    return texto or None


def consolidar_registros_importacao(registros: list[dict], avisos: list[str] | None = None) -> tuple[list[dict], list[str]]:
    """Consolida um lote para manter no máximo um lançamento efetivo por placa.

    A importação trabalha com um retrato operacional por data/turno. Se a mesma placa
    aparecer na frota e também em manutenção, a manutenção é priorizada por segurança,
    mas o contexto da linha operacional é preservado para revisão na prévia.
    """
    avisos_saida = list(avisos or [])
    por_placa: dict[str, list[dict]] = {}
    ordem: list[str] = []

    for registro in registros:
        item = dict(registro)
        placa = normalizar_placa_importacao(item.get("placa"))
        if not placa:
            continue
        item["placa"] = placa
        item.setdefault("alerta", None)
        if placa not in por_placa:
            por_placa[placa] = []
            ordem.append(placa)
        por_placa[placa].append(item)

    consolidados: list[dict] = []
    conflitos = 0
    duplicados = 0

    def preencher_vazios(destino: dict, origem: dict) -> None:
        for campo in ("tipo_veiculo", "motorista", "ajudante", "rota_id"):
            if not destino.get(campo) and origem.get(campo):
                destino[campo] = origem.get(campo)

    for placa in ordem:
        itens = por_placa[placa]
        if len(itens) == 1:
            consolidados.append(itens[0])
            continue

        manutencoes = [i for i in itens if normalizar_status_importacao(i.get("status")) == "MANUTENCAO" or str(i.get("tipo_registro") or "").upper() == "MANUTENCAO"]
        operacoes = [i for i in itens if i not in manutencoes]

        if manutencoes and operacoes:
            # Em um retrato operacional, manutenção é incompatível com circulação.
            # Mantemos uma única linha, porém sinalizamos a inconsistência para revisão.
            escolhido = dict(manutencoes[0])
            contexto = operacoes[0]
            preencher_vazios(escolhido, contexto)
            status_contexto = normalizar_status_importacao(contexto.get("status"))
            detalhes = []
            if contexto.get("rota_id"):
                detalhes.append(f"rota {contexto['rota_id']}")
            if contexto.get("motorista"):
                detalhes.append(f"motorista {contexto['motorista']}")
            complemento = f" ({', '.join(detalhes)})" if detalhes else ""
            escolhido["alerta"] = (
                f"Conflito: {placa} também aparece na frota como "
                f"{status_contexto.replace('_', ' ').title()}{complemento}. "
                "A manutenção foi priorizada; revise esta linha antes de confirmar."
            )
            confianca = escolhido.get("confianca")
            try:
                escolhido["confianca"] = min(float(confianca) if confianca is not None else 0.60, 0.60)
            except (TypeError, ValueError):
                escolhido["confianca"] = 0.60
            if not escolhido.get("origem_linha") and contexto.get("origem_linha"):
                escolhido["origem_linha"] = contexto.get("origem_linha")
            consolidados.append(escolhido)
            conflitos += 1
            duplicados += len(itens) - 1
            continue

        # Duplicatas do mesmo tipo: mantemos a melhor leitura e mesclamos campos vazios.
        escolhido = dict(max(
            itens,
            key=lambda i: float(i.get("confianca") or 0),
        ))
        statuses = {normalizar_status_importacao(i.get("status")) for i in itens}
        for outro in itens:
            preencher_vazios(escolhido, outro)
            if not escolhido.get("motivo") and outro.get("motivo"):
                escolhido["motivo"] = outro.get("motivo")
            if not escolhido.get("observacao") and outro.get("observacao"):
                escolhido["observacao"] = outro.get("observacao")

        if len(statuses) > 1:
            conflitos += 1
            escolhido["alerta"] = (
                f"Conflito: {placa} apareceu mais de uma vez com status diferentes: "
                + ", ".join(sorted(s.replace("_", " ").title() for s in statuses))
                + ". Revise a linha antes de confirmar."
            )
            escolhido["confianca"] = min(float(escolhido.get("confianca") or 0.70), 0.70)
        else:
            duplicados += len(itens) - 1
            escolhido["alerta"] = escolhido.get("alerta") or "Registro repetido consolidado automaticamente."
        consolidados.append(escolhido)

    if conflitos:
        avisos_saida.append(
            f"{conflitos} placa(s) aparecem com informações conflitantes. "
            "Essas linhas foram destacadas para revisão."
        )
    if duplicados:
        avisos_saida.append(
            f"{duplicados} ocorrência(s) duplicada(s) foram consolidadas para evitar lançamentos repetidos."
        )

    return consolidados, avisos_saida


def analisar_texto_operacional_local(texto: str) -> dict:
    registros = []
    avisos = []
    data_detectada = None
    unidade = None
    operador = None
    secao_manutencao = False

    match_data = re.search(r"\b(\d{2})/(\d{2})/(\d{4})\b", texto)
    if match_data:
        try:
            data_detectada = date(
                int(match_data.group(3)),
                int(match_data.group(2)),
                int(match_data.group(1)),
            )
        except ValueError:
            pass

    for linha_original in texto.splitlines():
        linha = linha_original.strip()
        if not linha:
            continue

        chave = criar_chave_busca(linha) or ""
        if chave.startswith("panorama "):
            unidade = linha.split(" ", 1)[1].strip() if " " in linha else None
            continue
        if chave.startswith("mlp:"):
            operador = linha.split(":", 1)[1].strip() if ":" in linha else None
            continue
        if "carros em manutencao" in chave:
            secao_manutencao = True
            continue
        if "frota fixa" in chave:
            secao_manutencao = False
            continue
        if (
            chave.startswith("quantidade ")
            or chave.startswith("legenda")
            or chave in {
                "carregando", "carro reserva/carregando", "indisponivel/motorista",
                "folga planejada motorista", "manutencao", "sem carga",
                "rodou em outro service"
            }
        ):
            continue

        match_placa = re.search(r"\b([A-Za-z]{3}[0-9][A-Za-z0-9][0-9]{2})\b", linha)
        if not match_placa:
            continue

        placa = normalizar_placa_importacao(match_placa.group(1))
        match_rota = re.search(r"\b(\d{8,10})\b", linha)
        rota_id = match_rota.group(1) if match_rota else None

        motorista = None
        parenteses = re.findall(r"\(([^()]*)\)", linha)
        if parenteses and not secao_manutencao:
            candidato = parenteses[-1].strip(" *")
            if candidato:
                motorista = candidato

        ajudante = None
        match_ajudante = re.search(r"(?:ajudante|aj)\s*[:\-]\s*([^|,;]+)", linha, flags=re.IGNORECASE)
        if match_ajudante:
            ajudante = match_ajudante.group(1).strip()

        status_linha, confianca = detectar_status_linha(linha, secao_manutencao)
        motivo = extrair_motivo_linha(linha, placa, rota_id, motorista)

        registros.append({
            "placa": placa,
            "tipo_registro": "MANUTENCAO" if status_linha == "MANUTENCAO" else "OPERACAO",
            "tipo_veiculo": None,
            "motorista": motorista,
            "ajudante": ajudante,
            "rota_id": rota_id,
            "status": status_linha,
            "motivo": motivo if status_linha in {"MANUTENCAO", "IMPEDIDO"} else None,
            "observacao": motivo if status_linha not in {"MANUTENCAO", "IMPEDIDO"} else None,
            "confianca": confianca,
            "origem_linha": linha,
            "alerta": None,
        })

    if not registros:
        avisos.append("Nenhum registro com placa foi identificado automaticamente.")
    elif any(item["status"] == "SEM_CLASSIFICACAO" for item in registros):
        avisos.append("Há registros sem status explícito. Revise-os antes de confirmar.")

    return {
        "data": data_detectada,
        "turno": None,
        "unidade": unidade,
        "operador": operador,
        "registros": registros,
        "avisos": avisos,
    }


def limpar_json_modelo(texto: str) -> dict:
    conteudo = texto.strip()
    if conteudo.startswith("```"):
        conteudo = re.sub(r"^```(?:json)?\s*", "", conteudo, flags=re.IGNORECASE)
        conteudo = re.sub(r"\s*```$", "", conteudo)
    inicio = conteudo.find("{")
    fim = conteudo.rfind("}")
    if inicio >= 0 and fim > inicio:
        conteudo = conteudo[inicio:fim + 1]
    return json.loads(conteudo)


def analisar_com_ia(texto: str | None = None, imagens: list[tuple[str, bytes]] | None = None) -> dict:
    if not OPENAI_API_KEY:
        raise HTTPException(
            status_code=503,
            detail=(
                "A análise de imagem por IA ainda não foi configurada. "
                "Defina OPENAI_API_KEY no .env do Ylume Ops."
            ),
        )

    try:
        from openai import OpenAI
    except ImportError as exc:
        raise HTTPException(
            status_code=503,
            detail="Instale a dependência openai para habilitar a análise por IA.",
        ) from exc

    instrucoes = """
Você é o extrator de dados operacionais do Ylume Ops.
Leia SOMENTE o conteúdo fornecido. Não invente placa, pessoa, rota, motivo ou status.
Devolva APENAS JSON válido, sem markdown, no formato:
{
  "data": "YYYY-MM-DD ou null",
  "turno": "Manhã, Tarde, Noite ou null",
  "unidade": "texto ou null",
  "operador": "texto ou null",
  "registros": [
    {
      "placa": "ABC1D23",
      "tipo_registro": "OPERACAO ou MANUTENCAO",
      "tipo_veiculo": null,
      "motorista": null,
      "ajudante": null,
      "rota_id": null,
      "status": "CARREGANDO|EM_ROTA|CONCLUIDA|RETORNANDO_ESTACAO|AMBULANCIA|RESERVA_CARREGANDO|FOLGA|IMPEDIDO|SEM_CARGA|OUTRO_SERVICE|INDISPONIVEL_MOTORISTA|SEM_CLASSIFICACAO|MANUTENCAO",
      "motivo": null,
      "observacao": null,
      "confianca": 0.0,
      "origem_linha": "trecho curto que originou o registro"
    }
  ],
  "avisos": []
}
Regras: normalize placas em maiúsculas; mantenha rota como texto; manutenção usa tipo_registro MANUTENCAO e status MANUTENCAO; se não houver evidência de status use SEM_CLASSIFICACAO; se não tiver certeza, reduza confianca; nunca deduza nomes ausentes.
""".strip()

    conteudo = [{"type": "input_text", "text": instrucoes}]
    if texto:
        conteudo.append({"type": "input_text", "text": "Conteúdo textual:\n" + texto})
    for mime, dados in imagens or []:
        b64 = base64.b64encode(dados).decode("ascii")
        conteudo.append({
            "type": "input_image",
            "image_url": f"data:{mime};base64,{b64}",
        })

    try:
        cliente = OpenAI(api_key=OPENAI_API_KEY)
        resposta = cliente.responses.create(
            model=YLUME_OPS_AI_MODEL,
            input=[{"role": "user", "content": conteudo}],
        )
        dados = limpar_json_modelo(resposta.output_text)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=502,
            detail="A IA não conseguiu analisar o conteúdo. Verifique a configuração do provedor e tente novamente.",
        ) from exc

    registros = []
    for item in dados.get("registros") or []:
        placa = normalizar_placa_importacao(str(item.get("placa") or ""))
        if not placa:
            continue
        status_item = normalizar_status_importacao(item.get("status"))
        tipo_registro = "MANUTENCAO" if status_item == "MANUTENCAO" else str(item.get("tipo_registro") or "OPERACAO").upper()
        if tipo_registro not in {"OPERACAO", "MANUTENCAO"}:
            tipo_registro = "OPERACAO"
        confianca = item.get("confianca")
        try:
            confianca = max(0.0, min(1.0, float(confianca))) if confianca is not None else None
        except (TypeError, ValueError):
            confianca = None
        registros.append({
            "placa": placa,
            "tipo_registro": tipo_registro,
            "tipo_veiculo": limpar_texto_opcional(item.get("tipo_veiculo")),
            "motorista": limpar_texto_opcional(item.get("motorista")),
            "ajudante": limpar_texto_opcional(item.get("ajudante")),
            "rota_id": limpar_texto_opcional(str(item.get("rota_id"))) if item.get("rota_id") is not None else None,
            "status": status_item,
            "motivo": limpar_texto_opcional(item.get("motivo")),
            "observacao": limpar_texto_opcional(item.get("observacao")),
            "confianca": confianca,
            "origem_linha": limpar_texto_opcional(item.get("origem_linha")),
            "alerta": limpar_texto_opcional(item.get("alerta")),
        })

    data_modelo = None
    if dados.get("data"):
        try:
            data_modelo = date.fromisoformat(str(dados["data"]))
        except ValueError:
            pass

    return {
        "data": data_modelo,
        "turno": limpar_texto_opcional(dados.get("turno")),
        "unidade": limpar_texto_opcional(dados.get("unidade")),
        "operador": limpar_texto_opcional(dados.get("operador")),
        "registros": registros,
        "avisos": [str(aviso) for aviso in (dados.get("avisos") or [])],
    }


def valor_coluna(linha: dict, *nomes: str) -> str | None:
    normalizado = {
        (criar_chave_busca(str(chave)) or ""): valor
        for chave, valor in linha.items()
    }
    for nome in nomes:
        valor = normalizado.get(criar_chave_busca(nome) or "")
        if valor is not None and str(valor).strip():
            return str(valor).strip()
    return None


def registros_de_linhas_estruturadas(linhas: list[dict]) -> list[dict]:
    registros = []
    for linha in linhas:
        placa = normalizar_placa_importacao(valor_coluna(linha, "placa", "veiculo", "veículo"))
        if not placa:
            continue
        status_item = normalizar_status_importacao(valor_coluna(linha, "status", "situacao", "situação", "classificacao", "classificação"))
        tipo_registro = "MANUTENCAO" if status_item == "MANUTENCAO" else "OPERACAO"
        registros.append({
            "placa": placa,
            "tipo_registro": tipo_registro,
            "tipo_veiculo": valor_coluna(linha, "tipo", "tipo veiculo", "tipo veículo"),
            "motorista": valor_coluna(linha, "motorista", "driver"),
            "ajudante": valor_coluna(linha, "ajudante", "helper"),
            "rota_id": valor_coluna(linha, "rota", "rota id", "route", "route id"),
            "status": status_item,
            "motivo": valor_coluna(linha, "motivo"),
            "observacao": valor_coluna(linha, "observacao", "observação", "obs"),
            "confianca": 1.0,
            "origem_linha": None,
            "alerta": None,
        })
    return registros


def obter_ou_criar_motorista_importacao(nome: str | None, db: Session, por_nome: dict) -> int | None:
    nome = limpar_texto_opcional(nome)
    if not nome:
        return None
    chave = criar_chave_busca(nome)
    motorista = por_nome.get(chave)
    if not motorista:
        motorista = models.Motorista(
            nome=nome,
            ativo=True,
            observacao="Cadastrado por importação inteligente.",
        )
        db.add(motorista)
        db.flush()
        por_nome[chave] = motorista
    if not motorista.ativo:
        return None
    return motorista.id


def obter_ou_criar_ajudante_importacao(nome: str | None, db: Session, por_nome: dict) -> int | None:
    nome = limpar_texto_opcional(nome)
    if not nome:
        return None
    chave = criar_chave_busca(nome)
    ajudante = por_nome.get(chave)
    if not ajudante:
        ajudante = models.Ajudante(
            nome=nome,
            ativo=True,
            observacao="Cadastrado por importação inteligente.",
        )
        db.add(ajudante)
        db.flush()
        por_nome[chave] = ajudante
    if not ajudante.ativo:
        return None
    return ajudante.id


# VEÍCULOS

@app.get(
    "/veiculos",
    response_model=list[schemas.VeiculoResponse],
    tags=["Veículos"],
)
def listar_veiculos(
    db: Session = Depends(get_db),
):
    return db.scalars(
        select(models.Veiculo)
        .order_by(models.Veiculo.placa)
    ).all()


@app.post(
    "/veiculos",
    response_model=schemas.VeiculoResponse,
    status_code=status.HTTP_201_CREATED,
    tags=["Veículos"],
)
def cadastrar_veiculo(
    veiculo: schemas.VeiculoCreate,
    db: Session = Depends(get_db),
):
    placa_normalizada = veiculo.placa.strip().upper()

    if not placa_normalizada:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Informe a placa do veículo.",
        )

    veiculo_existente = db.scalar(
        select(models.Veiculo)
        .where(models.Veiculo.placa == placa_normalizada)
    )

    if veiculo_existente:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Já existe um veículo cadastrado com esta placa.",
        )

    novo_veiculo = models.Veiculo(
        placa=placa_normalizada,
        tipo=limpar_texto_opcional(veiculo.tipo),
        categoria=veiculo.categoria.strip() or "Frota fixa",
        observacao=limpar_texto_opcional(veiculo.observacao),
        ativo=veiculo.ativo,
    )

    db.add(novo_veiculo)
    db.commit()
    db.refresh(novo_veiculo)

    return novo_veiculo


@app.patch(
    "/veiculos/{veiculo_id}",
    response_model=schemas.VeiculoResponse,
    tags=["Veículos"],
)
def atualizar_veiculo(
    veiculo_id: int,
    dados: schemas.VeiculoUpdate,
    db: Session = Depends(get_db),
):
    veiculo = db.get(models.Veiculo, veiculo_id)

    if not veiculo:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Veículo não encontrado.",
        )

    campos_enviados = dados.model_fields_set

    if "placa" in campos_enviados:
        if dados.placa is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="A placa não pode ficar vazia.",
            )

        nova_placa = dados.placa.strip().upper()

        if not nova_placa:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="A placa não pode ficar vazia.",
            )

        placa_em_uso = db.scalar(
            select(models.Veiculo)
            .where(
                models.Veiculo.placa == nova_placa,
                models.Veiculo.id != veiculo_id,
            )
        )

        if placa_em_uso:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Já existe outro veículo cadastrado com esta placa.",
            )

        veiculo.placa = nova_placa

    if "tipo" in campos_enviados:
        veiculo.tipo = limpar_texto_opcional(dados.tipo)

    if "categoria" in campos_enviados:
        if dados.categoria is None or not dados.categoria.strip():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="A categoria não pode ficar vazia.",
            )

        veiculo.categoria = dados.categoria.strip()

    if "observacao" in campos_enviados:
        veiculo.observacao = limpar_texto_opcional(dados.observacao)

    if "ativo" in campos_enviados and dados.ativo is not None:
        veiculo.ativo = dados.ativo

    db.commit()
    db.refresh(veiculo)

    return veiculo


@app.get(
    "/veiculos/{veiculo_id}/historico",
    response_model=schemas.VeiculoHistoricoResponse,
    tags=["Veículos"],
)
def obter_historico_veiculo(
    veiculo_id: int,
    db: Session = Depends(get_db),
):
    veiculo = db.get(models.Veiculo, veiculo_id)

    if not veiculo:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Veículo não encontrado.",
        )

    manutencoes = db.scalars(
        select(models.Manutencao)
        .where(models.Manutencao.veiculo_id == veiculo_id)
        .order_by(
            models.Manutencao.data_entrada.desc(),
            models.Manutencao.id.desc(),
        )
    ).all()

    operacoes = db.scalars(
        select(models.Operacao)
        .where(models.Operacao.veiculo_id == veiculo_id)
        .order_by(
            models.Operacao.data.desc(),
            models.Operacao.criado_em.desc(),
        )
    ).all()

    return {
        "veiculo": veiculo,
        "manutencoes": manutencoes,
        "operacoes": operacoes,
    }


@app.delete(
    "/veiculos/{veiculo_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    tags=["Veículos"],
)
def excluir_veiculo(
    veiculo_id: int,
    db: Session = Depends(get_db),
):
    veiculo = db.get(models.Veiculo, veiculo_id)

    if not veiculo:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Veículo não encontrado.",
        )

    possui_manutencao = db.scalar(
        select(models.Manutencao.id)
        .where(models.Manutencao.veiculo_id == veiculo_id)
        .limit(1)
    )

    possui_operacao = db.scalar(
        select(models.Operacao.id)
        .where(models.Operacao.veiculo_id == veiculo_id)
        .limit(1)
    )

    if possui_manutencao or possui_operacao:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                "Este veículo possui histórico e não pode ser excluído. "
                "Inative o veículo para mantê-lo fora da frota atual sem perder os registros."
            ),
        )

    db.delete(veiculo)
    db.commit()

    return None


# MANUTENÇÕES

@app.get(
    "/manutencoes",
    response_model=list[schemas.ManutencaoResponse],
    tags=["Manutenções"],
)
def listar_manutencoes(
    db: Session = Depends(get_db),
):
    return db.scalars(
        select(models.Manutencao)
        .order_by(
            models.Manutencao.data_entrada.desc(),
            models.Manutencao.id.desc(),
        )
    ).all()


@app.get(
    "/manutencoes/ativas",
    response_model=list[schemas.ManutencaoResponse],
    tags=["Manutenções"],
)
def listar_manutencoes_ativas(
    db: Session = Depends(get_db),
):
    return db.scalars(
        select(models.Manutencao)
        .where(models.Manutencao.status == "EM_MANUTENCAO")
        .order_by(
            models.Manutencao.data_entrada.desc(),
            models.Manutencao.id.desc(),
        )
    ).all()


@app.get(
    "/manutencoes/finalizadas",
    response_model=list[schemas.ManutencaoResponse],
    tags=["Manutenções"],
)
def listar_manutencoes_finalizadas(
    db: Session = Depends(get_db),
):
    return db.scalars(
        select(models.Manutencao)
        .where(models.Manutencao.status == "FINALIZADA")
        .order_by(
            models.Manutencao.data_retorno.desc(),
            models.Manutencao.id.desc(),
        )
    ).all()


@app.post(
    "/manutencoes",
    response_model=schemas.ManutencaoResponse,
    status_code=status.HTTP_201_CREATED,
    tags=["Manutenções"],
)
def cadastrar_manutencao(
    manutencao: schemas.ManutencaoCreate,
    db: Session = Depends(get_db),
):
    veiculo = db.get(models.Veiculo, manutencao.veiculo_id)

    if not veiculo:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Veículo não encontrado.",
        )

    if not veiculo.ativo:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Este veículo está inativo.",
        )

    motivo = manutencao.motivo.strip()

    if not motivo:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Informe o motivo da manutenção.",
        )

    if (
        manutencao.previsao_retorno is not None
        and manutencao.previsao_retorno < manutencao.data_entrada
    ):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="A previsão de retorno não pode ser anterior à data de entrada.",
        )

    manutencao_ativa = db.scalar(
        select(models.Manutencao)
        .where(
            models.Manutencao.veiculo_id == manutencao.veiculo_id,
            models.Manutencao.status == "EM_MANUTENCAO",
        )
    )

    if manutencao_ativa:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Este veículo já possui uma manutenção ativa.",
        )

    nova_manutencao = models.Manutencao(
        veiculo_id=manutencao.veiculo_id,
        motivo=motivo,
        data_entrada=manutencao.data_entrada,
        previsao_retorno=manutencao.previsao_retorno,
        status="EM_MANUTENCAO",
    )

    db.add(nova_manutencao)
    db.commit()
    db.refresh(nova_manutencao)

    return nova_manutencao


@app.patch(
    "/manutencoes/{manutencao_id}/finalizar",
    response_model=schemas.ManutencaoResponse,
    tags=["Manutenções"],
)
def finalizar_manutencao(
    manutencao_id: int,
    dados: schemas.ManutencaoFinalizar,
    db: Session = Depends(get_db),
):
    manutencao = db.get(models.Manutencao, manutencao_id)

    if not manutencao:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Manutenção não encontrada.",
        )

    if manutencao.status != "EM_MANUTENCAO":
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Esta manutenção já foi finalizada.",
        )

    data_retorno = dados.data_retorno or date.today()

    if data_retorno < manutencao.data_entrada:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="A data de retorno não pode ser anterior à data de entrada.",
        )

    manutencao.data_retorno = data_retorno
    manutencao.servico_realizado = limpar_texto_opcional(dados.servico_realizado)
    manutencao.condicao_retorno = limpar_texto_opcional(dados.condicao_retorno)
    manutencao.observacao_retorno = limpar_texto_opcional(dados.observacao_retorno)
    manutencao.oficina = limpar_texto_opcional(dados.oficina)
    manutencao.custo = dados.custo
    manutencao.status = "FINALIZADA"

    db.commit()
    db.refresh(manutencao)

    return manutencao


# MOTORISTAS

@app.get(
    "/motoristas",
    response_model=list[schemas.MotoristaResponse],
    tags=["Motoristas"],
)
def listar_motoristas(
    db: Session = Depends(get_db),
):
    return db.scalars(
        select(models.Motorista)
        .order_by(models.Motorista.nome)
    ).all()


@app.post(
    "/motoristas",
    response_model=schemas.MotoristaResponse,
    status_code=status.HTTP_201_CREATED,
    tags=["Motoristas"],
)
def cadastrar_motorista(
    motorista: schemas.MotoristaCreate,
    db: Session = Depends(get_db),
):
    nome_normalizado = motorista.nome.strip()

    if not nome_normalizado:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Informe o nome do motorista.",
        )

    cpf_normalizado = normalizar_cpf(motorista.cpf)

    if cpf_normalizado:
        cpf_em_uso = db.scalar(
            select(models.Motorista)
            .where(models.Motorista.cpf == cpf_normalizado)
        )

        if cpf_em_uso:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Já existe um motorista cadastrado com este CPF.",
            )

    novo_motorista = models.Motorista(
        nome=nome_normalizado,
        cpf=cpf_normalizado,
        telefone=limpar_texto_opcional(motorista.telefone),
        cnh=limpar_texto_opcional(motorista.cnh),
        categoria_cnh=limpar_texto_opcional(motorista.categoria_cnh),
        validade_cnh=motorista.validade_cnh,
        observacao=limpar_texto_opcional(motorista.observacao),
        ativo=motorista.ativo,
    )

    db.add(novo_motorista)
    db.commit()
    db.refresh(novo_motorista)

    return novo_motorista


@app.patch(
    "/motoristas/{motorista_id}",
    response_model=schemas.MotoristaResponse,
    tags=["Motoristas"],
)
def atualizar_motorista(
    motorista_id: int,
    dados: schemas.MotoristaUpdate,
    db: Session = Depends(get_db),
):
    motorista = db.get(models.Motorista, motorista_id)

    if not motorista:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Motorista não encontrado.",
        )

    campos_enviados = dados.model_fields_set

    if "nome" in campos_enviados:
        if dados.nome is None or not dados.nome.strip():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="O nome não pode ficar vazio.",
            )

        motorista.nome = dados.nome.strip()

    if "cpf" in campos_enviados:
        cpf_normalizado = normalizar_cpf(dados.cpf)

        if cpf_normalizado:
            cpf_em_uso = db.scalar(
                select(models.Motorista)
                .where(
                    models.Motorista.cpf == cpf_normalizado,
                    models.Motorista.id != motorista_id,
                )
            )

            if cpf_em_uso:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Já existe outro motorista cadastrado com este CPF.",
                )

        motorista.cpf = cpf_normalizado

    if "telefone" in campos_enviados:
        motorista.telefone = limpar_texto_opcional(dados.telefone)

    if "cnh" in campos_enviados:
        motorista.cnh = limpar_texto_opcional(dados.cnh)

    if "categoria_cnh" in campos_enviados:
        motorista.categoria_cnh = limpar_texto_opcional(dados.categoria_cnh)

    if "validade_cnh" in campos_enviados:
        motorista.validade_cnh = dados.validade_cnh

    if "observacao" in campos_enviados:
        motorista.observacao = limpar_texto_opcional(dados.observacao)

    if "ativo" in campos_enviados and dados.ativo is not None:
        motorista.ativo = dados.ativo

    db.commit()
    db.refresh(motorista)

    return motorista


@app.get(
    "/motoristas/{motorista_id}/historico",
    response_model=schemas.MotoristaHistoricoResponse,
    tags=["Motoristas"],
)
def obter_historico_motorista(
    motorista_id: int,
    db: Session = Depends(get_db),
):
    motorista = db.get(models.Motorista, motorista_id)

    if not motorista:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Motorista não encontrado.",
        )

    operacoes = db.scalars(
        select(models.Operacao)
        .where(models.Operacao.motorista_id == motorista_id)
        .order_by(
            models.Operacao.data.desc(),
            models.Operacao.criado_em.desc(),
        )
    ).all()

    return {
        "motorista": motorista,
        "operacoes": operacoes,
    }


@app.delete(
    "/motoristas/{motorista_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    tags=["Motoristas"],
)
def excluir_motorista(
    motorista_id: int,
    db: Session = Depends(get_db),
):
    motorista = db.get(models.Motorista, motorista_id)

    if not motorista:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Motorista não encontrado.",
        )

    possui_operacao = db.scalar(
        select(models.Operacao.id)
        .where(models.Operacao.motorista_id == motorista_id)
        .limit(1)
    )

    if possui_operacao:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                "Este motorista possui histórico e não pode ser excluído. "
                "Inative o motorista para mantê-lo fora da equipe atual sem perder os registros."
            ),
        )

    db.delete(motorista)
    db.commit()

    return None


# AJUDANTES

@app.get(
    "/ajudantes",
    response_model=list[schemas.AjudanteResponse],
    tags=["Ajudantes"],
)
def listar_ajudantes(db: Session = Depends(get_db)):
    return db.scalars(
        select(models.Ajudante).order_by(models.Ajudante.nome)
    ).all()


@app.post(
    "/ajudantes",
    response_model=schemas.AjudanteResponse,
    status_code=status.HTTP_201_CREATED,
    tags=["Ajudantes"],
)
def cadastrar_ajudante(
    ajudante: schemas.AjudanteCreate,
    db: Session = Depends(get_db),
):
    nome = ajudante.nome.strip()
    if not nome:
        raise HTTPException(status_code=400, detail="Informe o nome do ajudante.")

    cpf = normalizar_cpf(ajudante.cpf)
    if cpf and db.scalar(
        select(models.Ajudante).where(models.Ajudante.cpf == cpf)
    ):
        raise HTTPException(
            status_code=409,
            detail="Já existe um ajudante cadastrado com este CPF.",
        )

    novo = models.Ajudante(
        nome=nome,
        cpf=cpf,
        telefone=limpar_texto_opcional(ajudante.telefone),
        observacao=limpar_texto_opcional(ajudante.observacao),
        ativo=ajudante.ativo,
    )
    db.add(novo)
    db.commit()
    db.refresh(novo)
    return novo


@app.patch(
    "/ajudantes/{ajudante_id}",
    response_model=schemas.AjudanteResponse,
    tags=["Ajudantes"],
)
def atualizar_ajudante(
    ajudante_id: int,
    dados: schemas.AjudanteUpdate,
    db: Session = Depends(get_db),
):
    ajudante = db.get(models.Ajudante, ajudante_id)
    if not ajudante:
        raise HTTPException(status_code=404, detail="Ajudante não encontrado.")

    campos = dados.model_fields_set
    if "nome" in campos:
        if dados.nome is None or not dados.nome.strip():
            raise HTTPException(status_code=400, detail="O nome não pode ficar vazio.")
        ajudante.nome = dados.nome.strip()

    if "cpf" in campos:
        cpf = normalizar_cpf(dados.cpf)
        if cpf and not cpf_valido(cpf):
            raise HTTPException(status_code=400, detail="CPF inválido.")
        if cpf and db.scalar(
            select(models.Ajudante).where(
                models.Ajudante.cpf == cpf,
                models.Ajudante.id != ajudante_id,
            )
        ):
            raise HTTPException(
                status_code=409,
                detail="Já existe outro ajudante cadastrado com este CPF.",
            )
        ajudante.cpf = cpf

    if "telefone" in campos:
        ajudante.telefone = limpar_texto_opcional(dados.telefone)
    if "observacao" in campos:
        ajudante.observacao = limpar_texto_opcional(dados.observacao)
    if "ativo" in campos:
        ajudante.ativo = dados.ativo

    db.commit()
    db.refresh(ajudante)
    return ajudante


@app.delete(
    "/ajudantes/{ajudante_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    tags=["Ajudantes"],
)
def excluir_ajudante(
    ajudante_id: int,
    db: Session = Depends(get_db),
):
    ajudante = db.get(models.Ajudante, ajudante_id)
    if not ajudante:
        raise HTTPException(status_code=404, detail="Ajudante não encontrado.")

    possui_historico = db.scalar(
        select(models.Operacao.id)
        .where(models.Operacao.ajudante_id == ajudante_id)
        .limit(1)
    )
    if possui_historico:
        raise HTTPException(
            status_code=409,
            detail=(
                "Este ajudante possui histórico e não pode ser excluído. "
                "Arquive o cadastro para preservar os registros."
            ),
        )

    db.delete(ajudante)
    db.commit()
    return None


# OPERAÇÕES

STATUS_OPERACAO = {
    # STATUS OPERACIONAIS
    "CARREGANDO",
    "EM_ROTA",
    "CONCLUIDA",
    "RETORNANDO_ESTACAO",
    "AMBULANCIA",

    # STATUS MANUAIS DA OPERAÇÃO
    "RESERVA_CARREGANDO",
    "FOLGA",
    "IMPEDIDO",
    "SEM_CARGA",
    "OUTRO_SERVICE",
    "INDISPONIVEL_MOTORISTA",
    "SEM_CLASSIFICACAO",
}


@app.get(
    "/operacoes",
    response_model=list[schemas.OperacaoResponse],
    tags=["Operações"],
)
def listar_operacoes(
    data_operacao: date | None = None,
    turno: str | None = None,
    db: Session = Depends(get_db),
):
    consulta = select(models.Operacao)

    if data_operacao is not None:
        consulta = consulta.where(models.Operacao.data == data_operacao)

    if turno is not None:
        consulta = consulta.where(models.Operacao.turno == turno)

    consulta = consulta.order_by(models.Operacao.criado_em.desc())

    return db.scalars(consulta).all()


@app.post(
    "/operacoes",
    response_model=schemas.OperacaoResponse,
    status_code=status.HTTP_201_CREATED,
    tags=["Operações"],
)
def cadastrar_operacao(
    operacao: schemas.OperacaoCreate,
    db: Session = Depends(get_db),
):
    status_normalizado = operacao.status.strip().upper()
    turno_normalizado = operacao.turno.strip()

    if not turno_normalizado:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Informe o turno.",
        )

    if status_normalizado not in STATUS_OPERACAO:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Status de operação inválido.",
        )

    ajudante_id = None
    if operacao.ajudante_id is not None:
        ajudante = db.get(models.Ajudante, operacao.ajudante_id)
        if not ajudante:
            raise HTTPException(status_code=404, detail="Ajudante não encontrado.")
        if not ajudante.ativo:
            raise HTTPException(status_code=409, detail="Este ajudante está inativo.")
        ajudante_id = ajudante.id

    if operacao.veiculo_id is not None:
        veiculo = db.get(models.Veiculo, operacao.veiculo_id)

        if not veiculo:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Veículo não encontrado.",
            )

        if not veiculo.ativo:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Este veículo está inativo.",
            )

        manutencao_ativa = db.scalar(
            select(models.Manutencao)
            .where(
                models.Manutencao.veiculo_id == operacao.veiculo_id,
                models.Manutencao.status == "EM_MANUTENCAO",
            )
        )

        if manutencao_ativa:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=(
                    "Este veículo está em manutenção e não pode ser registrado na operação."
                ),
            )

        operacao_existente = db.scalar(
            select(models.Operacao)
            .where(
                models.Operacao.data == operacao.data,
                models.Operacao.turno == turno_normalizado,
                models.Operacao.veiculo_id == operacao.veiculo_id,
            )
        )

        if operacao_existente:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Este veículo já possui um registro neste turno.",
            )

    if operacao.motorista_id is not None:
        motorista = db.get(models.Motorista, operacao.motorista_id)

        if not motorista:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Motorista não encontrado.",
            )

        if not motorista.ativo:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Este motorista está inativo.",
            )

    nova_operacao = models.Operacao(
        data=operacao.data,
        turno=turno_normalizado,
        veiculo_id=operacao.veiculo_id,
        motorista_id=operacao.motorista_id,
        ajudante_id=ajudante_id,
        rota_id=limpar_texto_opcional(operacao.rota_id),
        status=status_normalizado,
        observacao=limpar_texto_opcional(operacao.observacao),
        origem=(operacao.origem.strip().upper() or "MANUAL"),
    )

    db.add(nova_operacao)
    db.commit()
    db.refresh(nova_operacao)

    return nova_operacao


@app.patch(
    "/operacoes/{operacao_id}",
    response_model=schemas.OperacaoResponse,
    tags=["Operações"],
)
def atualizar_operacao(
    operacao_id: int,
    dados: schemas.OperacaoUpdate,
    db: Session = Depends(get_db),
):
    operacao = db.get(models.Operacao, operacao_id)

    if not operacao:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Operação não encontrada.",
        )

    campos_enviados = dados.model_fields_set

    if "turno" in campos_enviados:
        if dados.turno is None or not dados.turno.strip():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="O turno não pode ficar vazio.",
            )

        operacao.turno = dados.turno.strip()

    if "veiculo_id" in campos_enviados:
        if dados.veiculo_id is None:
            operacao.veiculo_id = None
        else:
            veiculo = db.get(models.Veiculo, dados.veiculo_id)

            if not veiculo:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Veículo não encontrado.",
                )

            if not veiculo.ativo:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Este veículo está inativo.",
                )

            manutencao_ativa = db.scalar(
                select(models.Manutencao)
                .where(
                    models.Manutencao.veiculo_id == dados.veiculo_id,
                    models.Manutencao.status == "EM_MANUTENCAO",
                )
            )

            if manutencao_ativa:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Este veículo está em manutenção.",
                )

            operacao.veiculo_id = dados.veiculo_id

    if "motorista_id" in campos_enviados:
        if dados.motorista_id is None:
            operacao.motorista_id = None
        else:
            motorista = db.get(models.Motorista, dados.motorista_id)

            if not motorista:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Motorista não encontrado.",
                )

            if not motorista.ativo:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Este motorista está inativo.",
                )

            operacao.motorista_id = dados.motorista_id

    if "ajudante_id" in campos_enviados:
        if dados.ajudante_id is None:
            operacao.ajudante_id = None
        else:
            ajudante = db.get(models.Ajudante, dados.ajudante_id)
            if not ajudante:
                raise HTTPException(status_code=404, detail="Ajudante não encontrado.")
            if not ajudante.ativo:
                raise HTTPException(status_code=409, detail="Este ajudante está inativo.")
            operacao.ajudante_id = dados.ajudante_id

    if "rota_id" in campos_enviados:
        operacao.rota_id = limpar_texto_opcional(dados.rota_id)

    if "status" in campos_enviados:
        if dados.status is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="O status não pode ficar vazio.",
            )

        novo_status = dados.status.strip().upper()

        if novo_status not in STATUS_OPERACAO:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Status de operação inválido.",
            )

        operacao.status = novo_status

    if "observacao" in campos_enviados:
        operacao.observacao = limpar_texto_opcional(dados.observacao)

    if operacao.veiculo_id is not None:
        operacao_duplicada = db.scalar(
            select(models.Operacao)
            .where(
                models.Operacao.id != operacao.id,
                models.Operacao.data == operacao.data,
                models.Operacao.turno == operacao.turno,
                models.Operacao.veiculo_id == operacao.veiculo_id,
            )
        )

        if operacao_duplicada:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Este veículo já possui outro registro neste turno.",
            )

    db.commit()
    db.refresh(operacao)

    return operacao


@app.delete(
    "/operacoes/{operacao_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    tags=["Operações"],
)
def excluir_operacao(
    operacao_id: int,
    db: Session = Depends(get_db),
):
    operacao = db.get(models.Operacao, operacao_id)

    if not operacao:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Operação não encontrada.",
        )

    db.delete(operacao)
    db.commit()

    return None


# =====================================================
# IMPORTAÇÃO INTELIGENTE V6
# =====================================================

@app.get(
    "/importacoes/inteligente/status",
    tags=["Importações"],
)
def status_importacao_inteligente():
    return {
        "ia_configurada": bool(OPENAI_API_KEY),
        "modelo": YLUME_OPS_AI_MODEL if OPENAI_API_KEY else None,
        "formatos": ["texto", "csv", "xlsx", "png", "jpg", "jpeg", "webp"],
    }


@app.post(
    "/importacoes/inteligente/analisar-texto",
    response_model=schemas.ImportacaoInteligenteAnaliseResponse,
    tags=["Importações"],
)
def analisar_importacao_texto(
    dados: schemas.ImportacaoInteligenteTextoRequest,
):
    texto = dados.texto.strip()
    if not texto:
        raise HTTPException(status_code=400, detail="Cole algum conteúdo para analisar.")

    local = analisar_texto_operacional_local(texto)
    metodo = "parser_local"
    resultado = local

    # O parser local é preferido para o formato de panorama conhecido: não custa API
    # e mantém os dados dentro do servidor. A IA entra quando o formato não é reconhecido.
    if not local["registros"] and dados.usar_ia:
        resultado = analisar_com_ia(texto=texto)
        metodo = "ia"
    elif dados.usar_ia and OPENAI_API_KEY and any(
        item["status"] == "SEM_CLASSIFICACAO" for item in local["registros"]
    ):
        try:
            ia = analisar_com_ia(texto=texto)
            if len(ia["registros"]) >= len(local["registros"]):
                resultado = ia
                metodo = "ia"
        except HTTPException:
            # O parser local continua utilizável mesmo se a IA externa estiver indisponível.
            resultado = local
            resultado["avisos"].append(
                "A análise local foi usada porque o serviço de IA não respondeu."
            )

    if dados.data and not resultado.get("data"):
        resultado["data"] = dados.data
    if dados.turno and not resultado.get("turno"):
        resultado["turno"] = dados.turno.strip() or None

    resultado["registros"], resultado["avisos"] = consolidar_registros_importacao(
        resultado.get("registros") or [],
        resultado.get("avisos") or [],
    )

    return schemas.ImportacaoInteligenteAnaliseResponse(
        metodo=metodo,
        ia_configurada=bool(OPENAI_API_KEY),
        modelo=YLUME_OPS_AI_MODEL if metodo == "ia" else None,
        **resultado,
    )


@app.post(
    "/importacoes/inteligente/analisar-arquivos",
    response_model=schemas.ImportacaoInteligenteAnaliseResponse,
    tags=["Importações"],
)
async def analisar_importacao_arquivos(
    arquivos: list[UploadFile] = File(...),
    data_operacao: str = Form(""),
    turno: str = Form(""),
):
    if not arquivos:
        raise HTTPException(status_code=400, detail="Selecione ao menos um arquivo.")

    data_informada = None
    if data_operacao.strip():
        try:
            data_informada = date.fromisoformat(data_operacao.strip())
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Data inválida.") from exc

    imagens: list[tuple[str, bytes]] = []
    registros: list[dict] = []
    avisos: list[str] = []

    for arquivo in arquivos:
        conteudo = await arquivo.read()
        if len(conteudo) > MAX_IMPORT_FILE_BYTES:
            raise HTTPException(
                status_code=413,
                detail=f"O arquivo {arquivo.filename} excede o limite de 10 MB.",
            )

        nome = (arquivo.filename or "arquivo").lower()
        mime = (arquivo.content_type or "").lower()

        if mime in {"image/png", "image/jpeg", "image/webp"} or nome.endswith((".png", ".jpg", ".jpeg", ".webp")):
            if len(imagens) >= MAX_IMPORT_IMAGES:
                raise HTTPException(status_code=400, detail="Envie no máximo 4 imagens por análise.")
            mime_real = mime if mime.startswith("image/") else (
                "image/png" if nome.endswith(".png") else "image/jpeg"
            )
            imagens.append((mime_real, conteudo))
            continue

        if nome.endswith(".csv") or mime in {"text/csv", "application/csv"}:
            texto_csv = conteudo.decode("utf-8-sig", errors="replace")
            amostra = texto_csv[:4096]
            try:
                dialeto = csv.Sniffer().sniff(amostra, delimiters=",;\t|")
            except csv.Error:
                dialeto = csv.excel
                dialeto.delimiter = ";"
            leitor = csv.DictReader(io.StringIO(texto_csv), dialect=dialeto)
            registros.extend(registros_de_linhas_estruturadas(list(leitor)))
            continue

        if nome.endswith(".xlsx"):
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise HTTPException(
                    status_code=503,
                    detail="Instale openpyxl para importar arquivos Excel.",
                ) from exc
            workbook = load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
            planilha = workbook.active
            linhas = list(planilha.iter_rows(values_only=True))
            if linhas:
                cabecalhos = [str(valor or "").strip() for valor in linhas[0]]
                dicionarios = [dict(zip(cabecalhos, linha)) for linha in linhas[1:]]
                registros.extend(registros_de_linhas_estruturadas(dicionarios))
            workbook.close()
            continue

        if nome.endswith(".txt") or mime.startswith("text/"):
            resultado = analisar_texto_operacional_local(conteudo.decode("utf-8-sig", errors="replace"))
            registros.extend(resultado["registros"])
            avisos.extend(resultado["avisos"])
            continue

        avisos.append(f"Formato ignorado: {arquivo.filename}")

    if imagens:
        resultado_ia = analisar_com_ia(imagens=imagens)
        registros.extend(resultado_ia["registros"])
        avisos.extend(resultado_ia["avisos"])
        if not data_informada:
            data_informada = resultado_ia.get("data")
        if turno.strip() in {"", "Não informado", "Geral"} and resultado_ia.get("turno"):
            turno = resultado_ia.get("turno") or turno
        metodo = "ia_imagem" if not registros[:-len(resultado_ia["registros"])] else "misto"
    else:
        metodo = "arquivo_estruturado"

    registros, avisos = consolidar_registros_importacao(registros, avisos)

    if not registros:
        raise HTTPException(
            status_code=422,
            detail="Nenhum registro operacional foi identificado nos arquivos.",
        )

    return schemas.ImportacaoInteligenteAnaliseResponse(
        metodo=metodo,
        ia_configurada=bool(OPENAI_API_KEY),
        modelo=YLUME_OPS_AI_MODEL if imagens else None,
        data=data_informada,
        turno=turno.strip() or None,
        unidade=None,
        operador=None,
        registros=registros,
        avisos=avisos,
    )


@app.post(
    "/importacoes/inteligente/confirmar",
    response_model=schemas.ImportacaoInteligenteConfirmarResponse,
    tags=["Importações"],
)
def confirmar_importacao_inteligente(
    dados: schemas.ImportacaoInteligenteConfirmarRequest,
    db: Session = Depends(get_db),
):
    turno = dados.turno.strip()
    if not turno:
        raise HTTPException(status_code=400, detail="Informe o turno.")

    registros_consolidados_dict, _ = consolidar_registros_importacao(
        [registro.model_dump() for registro in dados.registros],
        [],
    )
    registros_confirmacao = [
        schemas.ImportacaoInteligenteRegistro(**registro)
        for registro in registros_consolidados_dict
    ]

    impedidos_sem_motivo = [
        normalizar_placa_importacao(registro.placa) or "sem placa"
        for registro in registros_confirmacao
        if normalizar_status_importacao(registro.status) == "IMPEDIDO"
        and not (limpar_texto_opcional(registro.motivo) or limpar_texto_opcional(registro.observacao))
    ]
    if impedidos_sem_motivo:
        amostra = ", ".join(impedidos_sem_motivo[:5])
        complemento = "" if len(impedidos_sem_motivo) <= 5 else f" e mais {len(impedidos_sem_motivo) - 5}"
        raise HTTPException(
            status_code=422,
            detail=f"Informe o motivo dos veículos impedidos antes de confirmar: {amostra}{complemento}.",
        )

    veiculos = db.scalars(select(models.Veiculo)).all()
    motoristas = db.scalars(select(models.Motorista)).all()
    ajudantes = db.scalars(select(models.Ajudante)).all()

    veiculos_por_placa = {normalizar_placa_importacao(v.placa): v for v in veiculos}
    motoristas_por_nome = {criar_chave_busca(m.nome): m for m in motoristas if criar_chave_busca(m.nome)}
    ajudantes_por_nome = {criar_chave_busca(a.nome): a for a in ajudantes if criar_chave_busca(a.nome)}

    op_importadas = op_atualizadas = manut_importadas = manut_atualizadas = ignorados = 0
    pendencias = []

    try:
        for registro in registros_confirmacao:
            placa = normalizar_placa_importacao(registro.placa)
            if not placa:
                ignorados += 1
                pendencias.append({"placa": None, "motivo": "Registro sem placa."})
                continue

            veiculo = veiculos_por_placa.get(placa)
            if not veiculo:
                veiculo = models.Veiculo(
                    placa=placa,
                    tipo=limpar_texto_opcional(registro.tipo_veiculo),
                    categoria="Importação inteligente",
                    observacao="Cadastrado automaticamente por importação inteligente.",
                    ativo=True,
                )
                db.add(veiculo)
                db.flush()
                veiculos_por_placa[placa] = veiculo
            elif registro.tipo_veiculo and not limpar_texto_opcional(veiculo.tipo):
                veiculo.tipo = limpar_texto_opcional(registro.tipo_veiculo)

            if not veiculo.ativo:
                ignorados += 1
                pendencias.append({"placa": placa, "motivo": "Veículo inativo. Reative-o antes de importar."})
                continue

            status_item = normalizar_status_importacao(registro.status)
            tipo_registro = registro.tipo_registro.strip().upper()
            if status_item == "MANUTENCAO":
                tipo_registro = "MANUTENCAO"

            if tipo_registro == "MANUTENCAO":
                manutencao = db.scalar(
                    select(models.Manutencao).where(
                        models.Manutencao.veiculo_id == veiculo.id,
                        models.Manutencao.status == "EM_MANUTENCAO",
                    ).order_by(models.Manutencao.data_entrada.desc())
                )
                motivo = limpar_texto_opcional(registro.motivo) or limpar_texto_opcional(registro.observacao) or "Manutenção informada na importação."
                if manutencao:
                    if registro.motivo:
                        manutencao.motivo = motivo
                    manut_atualizadas += 1
                else:
                    db.add(models.Manutencao(
                        veiculo_id=veiculo.id,
                        motivo=motivo,
                        data_entrada=dados.data,
                        status="EM_MANUTENCAO",
                    ))
                    manut_importadas += 1
                continue

            if status_item not in STATUS_OPERACAO:
                ignorados += 1
                pendencias.append({"placa": placa, "motivo": f"Status não reconhecido: {registro.status}"})
                continue

            manutencao_periodo = db.scalar(
                select(models.Manutencao.id).where(
                    models.Manutencao.veiculo_id == veiculo.id,
                    models.Manutencao.data_entrada <= dados.data,
                    or_(
                        models.Manutencao.data_retorno.is_(None),
                        models.Manutencao.data_retorno >= dados.data,
                    ),
                ).limit(1)
            )
            if manutencao_periodo:
                ignorados += 1
                pendencias.append({"placa": placa, "motivo": "Veículo consta em manutenção na data selecionada."})
                continue

            motorista_id = obter_ou_criar_motorista_importacao(registro.motorista, db, motoristas_por_nome)
            if registro.motorista and motorista_id is None:
                ignorados += 1
                pendencias.append({"placa": placa, "motivo": f"Motorista inativo: {registro.motorista}."})
                continue

            ajudante_id = obter_ou_criar_ajudante_importacao(registro.ajudante, db, ajudantes_por_nome)
            if registro.ajudante and ajudante_id is None:
                ignorados += 1
                pendencias.append({"placa": placa, "motivo": f"Ajudante inativo: {registro.ajudante}."})
                continue

            operacao = db.scalar(
                select(models.Operacao).where(
                    models.Operacao.data == dados.data,
                    models.Operacao.turno == turno,
                    models.Operacao.veiculo_id == veiculo.id,
                )
            )
            observacao = limpar_texto_opcional(registro.observacao) or limpar_texto_opcional(registro.motivo)

            if operacao:
                if (operacao.origem or "").upper() == "MANUAL" and not dados.sobrescrever_manuais:
                    ignorados += 1
                    pendencias.append({
                        "placa": placa,
                        "motivo": "Já existe registro manual. Ele foi preservado.",
                    })
                    continue
                operacao.motorista_id = motorista_id
                operacao.ajudante_id = ajudante_id
                operacao.rota_id = limpar_texto_opcional(registro.rota_id)
                operacao.status = status_item
                operacao.observacao = observacao
                operacao.origem = dados.origem.strip().upper() or "IMPORTACAO_INTELIGENTE"
                op_atualizadas += 1
            else:
                db.add(models.Operacao(
                    data=dados.data,
                    turno=turno,
                    veiculo_id=veiculo.id,
                    motorista_id=motorista_id,
                    ajudante_id=ajudante_id,
                    rota_id=limpar_texto_opcional(registro.rota_id),
                    status=status_item,
                    observacao=observacao,
                    origem=dados.origem.strip().upper() or "IMPORTACAO_INTELIGENTE",
                ))
                op_importadas += 1

        db.commit()
    except Exception:
        db.rollback()
        raise

    return schemas.ImportacaoInteligenteConfirmarResponse(
        recebidos=len(registros_confirmacao),
        operacoes_importadas=op_importadas,
        operacoes_atualizadas=op_atualizadas,
        manutencoes_importadas=manut_importadas,
        manutencoes_atualizadas=manut_atualizadas,
        ignorados=ignorados,
        pendencias=pendencias,
    )


# IMPORTAÇÃO ASSISTIDA

@app.post(
    "/importacoes/operacao",
    response_model=schemas.ImportacaoOperacaoResponse,
    tags=["Importações"],
)
def importar_operacao(
    importacao: schemas.ImportacaoOperacaoRequest,
    db: Session = Depends(get_db),
):
    turno = importacao.turno.strip()

    if not turno:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Informe o turno da operação.",
        )

    origem = (
        importacao.origem.strip().upper()
        or
        "IMPORTACAO_ASSISTIDA"
    )

    veiculos = db.scalars(
        select(models.Veiculo)
    ).all()

    motoristas = db.scalars(
        select(models.Motorista)
    ).all()

    veiculos_por_placa = {
        veiculo.placa.strip().upper(): veiculo
        for veiculo in veiculos
    }

    motoristas_por_nome = {
        criar_chave_busca(
            motorista.nome
        ): motorista
        for motorista in motoristas
        if criar_chave_busca(
            motorista.nome
        )
    }

    manutencoes_no_periodo = db.scalars(
        select(models.Manutencao)
        .where(
            models.Manutencao.data_entrada
            <=
            importacao.data,
            or_(
                models.Manutencao.data_retorno.is_(None),
                models.Manutencao.data_retorno
                >=
                importacao.data,
            ),
        )
    ).all()

    ids_em_manutencao = {
        manutencao.veiculo_id
        for manutencao in manutencoes_no_periodo
    }

    importados = 0
    atualizados = 0
    ignorados = 0
    pendencias = []

    try:
        for registro in importacao.registros:
            placa = (
                registro.placa
                .strip()
                .upper()
            )

            status_registro = (
                registro.status
                .strip()
                .upper()
            )

            if not placa:
                ignorados += 1

                pendencias.append({
                    "placa": None,
                    "motivo": "Registro sem placa.",
                })

                continue

            if (
                status_registro
                not in
                STATUS_OPERACAO
            ):
                ignorados += 1

                pendencias.append({
                    "placa": placa,
                    "motivo": (
                        "Status não reconhecido: "
                        f"{status_registro}"
                    ),
                })

                continue

            tipo_veiculo = (
                limpar_texto_opcional(
                    registro.tipo_veiculo
                )
            )

            veiculo = (
                veiculos_por_placa.get(
                    placa
                )
            )

            if not veiculo:
                veiculo = models.Veiculo(
                    placa=placa,
                    tipo=tipo_veiculo,
                    categoria="Importação assistida",
                    observacao=(
                        "Cadastrado automaticamente "
                        "por importação assistida."
                    ),
                    ativo=True,
                )

                db.add(
                    veiculo
                )

                db.flush()

                veiculos_por_placa[
                    placa
                ] = veiculo

            elif tipo_veiculo:
                tipo_atual = (
                    limpar_texto_opcional(
                        veiculo.tipo
                    )
                )

                cadastro_automatico = (
                    veiculo.categoria
                    ==
                    "Importação assistida"
                )

                if (
                    not tipo_atual
                    or
                    cadastro_automatico
                ):
                    veiculo.tipo = (
                        tipo_veiculo
                    )

            if not veiculo.ativo:
                ignorados += 1

                pendencias.append({
                    "placa": placa,
                    "motivo": (
                        "Veículo está inativo. "
                        "O cadastro não foi reativado "
                        "automaticamente."
                    ),
                })

                continue

            if (
                veiculo.id
                in
                ids_em_manutencao
            ):
                ignorados += 1

                pendencias.append({
                    "placa": placa,
                    "motivo": (
                        "Veículo estava em manutenção "
                        "na data consultada."
                    ),
                })

                continue

            motorista_id = None

            if registro.motorista:
                nome_motorista = (
                    registro.motorista
                    .strip()
                )

                chave_motorista = (
                    criar_chave_busca(
                        nome_motorista
                    )
                )

                motorista = (
                    motoristas_por_nome.get(
                        chave_motorista
                    )
                    if chave_motorista
                    else None
                )

                if not motorista:
                    motorista = models.Motorista(
                        nome=nome_motorista,
                        telefone=None,
                        observacao=(
                            "Cadastrado automaticamente "
                            "por importação assistida."
                        ),
                        ativo=True,
                    )

                    db.add(
                        motorista
                    )

                    db.flush()

                    if chave_motorista:
                        motoristas_por_nome[
                            chave_motorista
                        ] = motorista

                if not motorista.ativo:
                    ignorados += 1

                    pendencias.append({
                        "placa": placa,
                        "motivo": (
                            "Motorista está inativo: "
                            f"{nome_motorista}. "
                            "O cadastro não foi reativado "
                            "automaticamente."
                        ),
                    })

                    continue

                motorista_id = (
                    motorista.id
                )

            operacao_existente = db.scalar(
                select(models.Operacao)
                .where(
                    models.Operacao.data
                    ==
                    importacao.data,
                    models.Operacao.turno
                    ==
                    turno,
                    models.Operacao.veiculo_id
                    ==
                    veiculo.id,
                )
            )

            rota_id = (
                limpar_texto_opcional(
                    registro.rota_id
                )
            )

            observacao = (
                combinar_textos_operacao(
                    registro.cluster,
                    registro.observacao,
                )
            )

            if operacao_existente:
                origem_existente = (
                    operacao_existente.origem
                    or
                    ""
                ).strip().upper()

                if origem_existente == "MANUAL":
                    ignorados += 1

                    pendencias.append({
                        "placa": placa,
                        "motivo": (
                            "Já existe um registro manual "
                            "neste turno. O registro manual "
                            "foi preservado."
                        ),
                    })

                    continue

                operacao_existente.motorista_id = (
                    motorista_id
                )

                operacao_existente.rota_id = (
                    rota_id
                )

                operacao_existente.status = (
                    status_registro
                )

                operacao_existente.observacao = (
                    observacao
                )

                operacao_existente.origem = (
                    origem
                )

                atualizados += 1

            else:
                nova_operacao = models.Operacao(
                    data=importacao.data,
                    turno=turno,
                    veiculo_id=veiculo.id,
                    motorista_id=motorista_id,
                    rota_id=rota_id,
                    status=status_registro,
                    observacao=observacao,
                    origem=origem,
                )

                db.add(
                    nova_operacao
                )

                importados += 1

        db.commit()

    except Exception:
        db.rollback()
        raise

    return schemas.ImportacaoOperacaoResponse(
        recebidos=len(
            importacao.registros
        ),
        importados=importados,
        atualizados=atualizados,
        ignorados=ignorados,
        pendencias=pendencias,
    )


# =====================================================
# VEÍCULOS SEM REGISTRO NO TURNO
# =====================================================

@app.get(
    "/operacoes/veiculos-sem-registro",
    response_model=list[
        schemas.VeiculoSemRegistroResponse
    ],
    tags=["Operações"],
)
def listar_veiculos_sem_registro(
    data_operacao: date,
    turno: str,
    db: Session = Depends(get_db),
):
    turno_normalizado = turno.strip()

    if not turno_normalizado:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Informe o turno.",
        )

    veiculos = db.scalars(
        select(models.Veiculo)
        .where(
            models.Veiculo.ativo.is_(True)
        )
        .order_by(
            models.Veiculo.placa
        )
    ).all()

    operacoes = db.scalars(
        select(models.Operacao)
        .where(
            models.Operacao.data
            ==
            data_operacao,
            models.Operacao.turno
            ==
            turno_normalizado,
        )
    ).all()

    manutencoes = db.scalars(
        select(models.Manutencao)
        .where(
            models.Manutencao.data_entrada
            <=
            data_operacao,
            or_(
                models.Manutencao.data_retorno.is_(None),
                models.Manutencao.data_retorno
                >=
                data_operacao,
            ),
        )
    ).all()

    ids_com_operacao = {
        operacao.veiculo_id
        for operacao in operacoes
        if operacao.veiculo_id is not None
    }

    ids_em_manutencao = {
        manutencao.veiculo_id
        for manutencao in manutencoes
    }

    return [
        veiculo
        for veiculo in veiculos
        if (
            veiculo.id
            not in
            ids_com_operacao
            and
            veiculo.id
            not in
            ids_em_manutencao
        )
    ]


CLASSIFICACOES_AUSENCIA = {
    "MANUTENCAO",
    "FOLGA",
    "IMPEDIDO",
    "SEM_CARGA",
    "OUTRO_SERVICE",
    "INDISPONIVEL_MOTORISTA",
}


@app.post(
    "/operacoes/veiculos/{veiculo_id}/classificar",
    response_model=(
        schemas.ClassificarVeiculoAusenteResponse
    ),
    tags=["Operações"],
)
def classificar_veiculo_ausente(
    veiculo_id: int,
    dados: (
        schemas.ClassificarVeiculoAusenteRequest
    ),
    db: Session = Depends(get_db),
):
    veiculo = db.get(
        models.Veiculo,
        veiculo_id,
    )

    if not veiculo:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Veículo não encontrado.",
        )

    if not veiculo.ativo:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Este veículo está inativo.",
        )

    turno = dados.turno.strip()

    if not turno:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Informe o turno.",
        )

    classificacao = (
        dados.classificacao
        .strip()
        .upper()
    )

    if (
        classificacao
        not in
        CLASSIFICACOES_AUSENCIA
    ):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "Classificação inválida. Use "
                "MANUTENCAO, FOLGA, IMPEDIDO, "
                "SEM_CARGA, OUTRO_SERVICE ou "
                "INDISPONIVEL_MOTORISTA."
            ),
        )

    operacao_existente = db.scalar(
        select(models.Operacao)
        .where(
            models.Operacao.data
            ==
            dados.data,
            models.Operacao.turno
            ==
            turno,
            models.Operacao.veiculo_id
            ==
            veiculo.id,
        )
    )

    if operacao_existente:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                "Este veículo já possui um registro "
                "neste turno."
            ),
        )

    motivo = (
        limpar_texto_opcional(
            dados.motivo
        )
    )

    if classificacao == "MANUTENCAO":
        manutencao_existente = db.scalar(
            select(models.Manutencao)
            .where(
                models.Manutencao.veiculo_id
                ==
                veiculo.id,
                models.Manutencao.data_entrada
                <=
                dados.data,
                or_(
                    models.Manutencao.data_retorno.is_(
                        None
                    ),
                    models.Manutencao.data_retorno
                    >=
                    dados.data,
                ),
            )
        )

        if manutencao_existente:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=(
                    "O veículo já possui manutenção "
                    "registrada nesta data."
                ),
            )

        motivo_manutencao = (
            motivo
            or
            "Manutenção informada pela operação."
        )

        nova_manutencao = models.Manutencao(
            veiculo_id=veiculo.id,
            motivo=motivo_manutencao,
            data_entrada=dados.data,
            previsao_retorno=(
                dados.previsao_retorno
            ),
            status="EM_MANUTENCAO",
        )

        db.add(
            nova_manutencao
        )

        db.commit()
        db.refresh(
            nova_manutencao
        )

        return {
            "tipo_registro": "MANUTENCAO",
            "mensagem": (
                f"{veiculo.placa} foi registrado "
                "em manutenção."
            ),
            "operacao": None,
            "manutencao": nova_manutencao,
        }

    nova_operacao = models.Operacao(
        data=dados.data,
        turno=turno,
        veiculo_id=veiculo.id,
        motorista_id=None,
        rota_id=None,
        status=classificacao,
        observacao=motivo,
        origem="MANUAL",
    )

    db.add(
        nova_operacao
    )

    db.commit()
    db.refresh(
        nova_operacao
    )

    return {
        "tipo_registro": "OPERACAO",
        "mensagem": (
            f"{veiculo.placa} foi classificado "
            "com sucesso."
        ),
        "operacao": nova_operacao,
        "manutencao": None,
    }

# =====================================================
# PANORAMA
# =====================================================

LEGENDA_STATUS = {
    "CARREGANDO": "✅",
    "EM_ROTA": "✅",
    "CONCLUIDA": "✅",
    "RETORNANDO_ESTACAO": "🔄",
    "AMBULANCIA": "🚑",
    "RESERVA_CARREGANDO": "🚗",
    "FOLGA": "⚠️",
    "IMPEDIDO": "🚫",
    "SEM_CARGA": "📦",
    "OUTRO_SERVICE": "🔄",
    "INDISPONIVEL_MOTORISTA": "⏸️",
    "SEM_CLASSIFICACAO": "",
}


def obter_configuracao_panorama(db: Session) -> models.PanoramaConfiguracao:
    config = db.get(models.PanoramaConfiguracao, 1)
    if not config:
        config = models.PanoramaConfiguracao(
            id=1,
            unidade="Base operacional",
            operador="",
        )
        db.add(config)
        db.commit()
        db.refresh(config)
    return config


@app.get(
    "/configuracao-panorama",
    response_model=schemas.PanoramaConfiguracaoResponse,
    tags=["Panorama"],
)
def ler_configuracao_panorama(db: Session = Depends(get_db)):
    return obter_configuracao_panorama(db)


@app.put(
    "/configuracao-panorama",
    response_model=schemas.PanoramaConfiguracaoResponse,
    tags=["Panorama"],
)
def salvar_configuracao_panorama(
    dados: schemas.PanoramaConfiguracaoUpdate,
    db: Session = Depends(get_db),
):
    config = obter_configuracao_panorama(db)
    if dados.unidade is not None:
        config.unidade = dados.unidade.strip() or "Base operacional"
    if dados.operador is not None:
        config.operador = dados.operador.strip()
    db.commit()
    db.refresh(config)
    return config


@app.get(
    "/panorama",
    response_model=schemas.PanoramaResponse,
    tags=["Panorama"],
)
def gerar_panorama(
    data_operacao: date,
    turno: str | None = None,
    db: Session = Depends(get_db),
):
    config = obter_configuracao_panorama(db)
    veiculos = db.scalars(
        select(models.Veiculo)
        .where(models.Veiculo.ativo.is_(True))
        .order_by(models.Veiculo.placa)
    ).all()

    manutencoes = db.scalars(
        select(models.Manutencao).where(
            models.Manutencao.data_entrada <= data_operacao,
            or_(
                models.Manutencao.data_retorno.is_(None),
                models.Manutencao.data_retorno >= data_operacao,
            ),
        )
    ).all()

    consulta = select(models.Operacao).where(
        models.Operacao.data == data_operacao
    )
    if turno:
        consulta = consulta.where(models.Operacao.turno == turno)
    operacoes = db.scalars(
        consulta.order_by(models.Operacao.criado_em)
    ).all()

    motoristas = {
        item.id: item
        for item in db.scalars(select(models.Motorista)).all()
    }
    ajudantes = {
        item.id: item
        for item in db.scalars(select(models.Ajudante)).all()
    }
    veiculos_por_id = {item.id: item for item in veiculos}
    operacao_por_veiculo = {
        item.veiculo_id: item
        for item in operacoes
        if item.veiculo_id is not None
    }
    ids_manutencao = {item.veiculo_id for item in manutencoes}

    ociosos = [
        item
        for item in veiculos
        if item.id not in operacao_por_veiculo
        and item.id not in ids_manutencao
    ]
    ociosos.extend(
        veiculos_por_id[item.veiculo_id]
        for item in operacoes
        if item.veiculo_id in veiculos_por_id
        and item.status in {"SEM_CLASSIFICACAO", "INDISPONIVEL_MOTORISTA"}
    )
    vistos = set()
    ociosos = [
        item for item in ociosos
        if not (item.id in vistos or vistos.add(item.id))
    ]

    linhas = [
        f"PANORAMA {config.unidade}",
        f"MLP: {config.operador}" if config.operador else "MLP:",
        data_operacao.strftime("%d/%m/%Y"),
        "",
        f"Quantidade total de veiculos na base: {len(veiculos)}",
        f"Quantidade total de veiculos em manutenção: {len(manutencoes)}",
        f"Quantidade de veiculos ociosos: {len(ociosos)}",
        "",
        "Legenda:",
        "",
        "✅ Carregando",
        "🚗 Carro reserva/Carregando",
        "⏸️ Indisponível/motorista",
        "⚠️ Folga planejada motorista",
        "🛠️ Manutenção",
        "🚫 Impedido de rodar no dia. (informar o motivo)",
        "📦 Sem carga",
        "🔄 Rodou em outro service",
        "",
        "FROTA FIXA",
        "",
    ]

    for veiculo in veiculos:
        operacao = operacao_por_veiculo.get(veiculo.id)
        if not operacao:
            if veiculo.id not in ids_manutencao:
                linhas.append(veiculo.placa)
            continue

        partes = [veiculo.placa]
        emoji = LEGENDA_STATUS.get(operacao.status, "")
        if emoji:
            partes.append(emoji)
        if operacao.status == "FOLGA":
            partes.append("Folga")
        if operacao.rota_id:
            partes.append(operacao.rota_id)

        motorista = motoristas.get(operacao.motorista_id)
        if motorista:
            partes.append(f"({motorista.nome})")

        ajudante = ajudantes.get(operacao.ajudante_id)
        if ajudante:
            partes.append(f"[Ajudante: {ajudante.nome}]")

        if operacao.observacao:
            partes.append(operacao.observacao)

        linhas.append(" ".join(partes))

    linhas.extend(["", "CARROS EM MANUTENÇÃO.", ""])

    for manutencao in sorted(
        manutencoes,
        key=lambda item: (
            veiculos_por_id.get(item.veiculo_id).placa
            if item.veiculo_id in veiculos_por_id
            else ""
        ),
    ):
        veiculo = veiculos_por_id.get(manutencao.veiculo_id)
        if not veiculo:
            continue
        motivo = (manutencao.motivo or "").strip()
        linhas.append(
            f"{veiculo.placa} 🛠️" + (f" {motivo}" if motivo else "")
        )

    return schemas.PanoramaResponse(
        data=data_operacao,
        turno=turno,
        total_veiculos=len(veiculos),
        veiculos_manutencao=len(manutencoes),
        veiculos_operacao=len(operacoes),
        veiculos_ociosos=len(ociosos),
        texto="\n".join(linhas),
    )
